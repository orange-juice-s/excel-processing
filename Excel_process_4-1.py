# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from heapq import nlargest, nsmallest
from numpy import mean
from openpyxl.chart import Reference, LineChart
from pandas import read_excel, read_csv
import os.path

# 常量定义
ASTERISK = 100
B2636_START_COL = 4
PDA_START_COL = 2
B2636_PERCENTAGE_IGNORED = 0.2
PDA_PERCENTAGE_IGNORED = 0.35
PHOTO_CURRENT_COUNT_NUMBERS = 20


# 创建目标文件夹 ./xxx_processed
def mkdir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")

    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False


# 处理目录，获取源目录下的文件目录列表和目标文件夹下的文件目录列表
class ExcelDir(object):

    def __init__(self, SourcePath: str, TestDevice: str):
        self.source_path = SourcePath
        self.target_path = os.path.dirname(self.source_path) + "_processed/"
        self.test_device = TestDevice

    @staticmethod
    def IsSubString(SubStrList, Str):
        """
        #判断字符串Str是否包含序列SubStrList中的每一个子字符串
        #>>>SubStrList=['F','EMS','txt']
        #>>>Str='F06925EMS91.txt'
        #>>>IsSubString(SubStrList,Str)#return True (or False)
        :param SubStrList:
        :param Str:
        :return:
        """
        flag = True
        for substr in SubStrList:
            if not (substr in Str):
                flag = False
        return flag

    def SourceFileList(self):
        """
        #获取目录中指定的文件名
        #>>>FlagStr=['F','EMS','txt'] #要求文件名称中包含这些字符
        #>>>source_file_list=GetFileList(FindPath,FlagStr) #
        """

        FlagStr = ".xlsx"

        from os import listdir, path
        source_file_list = []
        FileNames = listdir(self.source_path)
        if len(FileNames) > 0:
            for fn in FileNames:
                if len(FlagStr) > 0:
                    # 返回指定类型的文件名
                    if self.IsSubString(FlagStr, fn):
                        full_file_name = path.join(self.source_path, fn)
                        source_file_list.append(full_file_name)
                else:
                    # 默认直接返回所有文件名
                    full_file_name = path.join(self.source_path, fn)
                    source_file_list.append(full_file_name)

        # 对文件名排序
        if len(source_file_list) > 0:
            source_file_list.sort()

        return source_file_list

    def TargetFileList(self):

        source_file_list = self.SourceFileList()
        target_file_list = []
        for dirs in source_file_list:
            directory = dirs.replace(self.source_path, self.target_path, 1)
            target_file_list.append(directory)
        return target_file_list


# 处理Excel
class ExcelProcess(ExcelDir):

    def __init__(self, SourcePath, TestDevice, SourceFileName, TargetFileName):
        super().__init__(SourcePath, TestDevice)
        # 获取源文件目录的列表 不能在这里调用 for 否则只能读出列表最后一项的函数值
        self.target_file_name = TargetFileName
        self.source_file_name = SourceFileName
        self.wb = load_workbook(SourceFileName)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        # 判断测试设备
        if self.test_device == '2636B':
            self.start_row = int(self.ws.max_row * B2636_PERCENTAGE_IGNORED)
            self.start_col = B2636_START_COL
        elif self.test_device == 'PDA':
            self.start_row = int(self.ws.max_row * PDA_PERCENTAGE_IGNORED)
            self.start_col = PDA_START_COL
        else:
            pass

    # 定义一个获取前20个最大最小值的平均数（去除了表格前10-15%的抖动区域），即获取光电流和暗电流
    def getAverageValue(self):
        print("-" * ASTERISK)
        print("正在获取 %s 的光暗电流值" % self.source_file_name)

        if self.test_device == "PDA":
            data = read_excel(self.source_file_name, header=None, usecols=[1], skiprows=self.start_row)
            data_abs = data.abs()
            column_data = data_abs[1].to_list()
            column_max = nlargest(PHOTO_CURRENT_COUNT_NUMBERS, column_data)
            column_min = nsmallest(PHOTO_CURRENT_COUNT_NUMBERS, column_data)
            # 取其平均值
            averageMaxValue = mean(column_max)
            averageMinValue = mean(column_min)
            # 返回这两个平均值（光电流和暗电流）
            return averageMaxValue, averageMinValue
        else:
            data = read_excel(self.source_file_name, header=None, usecols=[3], skiprows=self.start_row)
            data_abs = data.abs()
            column_data = data_abs[3].to_list()
            column_max = nlargest(PHOTO_CURRENT_COUNT_NUMBERS, column_data)
            column_min = nsmallest(PHOTO_CURRENT_COUNT_NUMBERS, column_data)
            # 取其平均值
            averageMaxValue = mean(column_max)
            averageMinValue = mean(column_min)
            # 返回这两个平均值（光电流和暗电流）
            return averageMaxValue, averageMinValue

    # 在新的path中画一个简要的I-t曲线。x轴是1， 2， 3， ... 未用到真实的t时间。
    def re_draw(self):

        c1 = LineChart()
        c1.title = "2D Line Chart"
        c1.legend = None
        c1.style = 15
        c1.y_axis.title = 'Size'
        c1.x_axis.title = 'Test Number'

        data = Reference(self.ws, min_col=self.start_col, min_row=self.start_row, max_col=self.start_col,
                         max_row=self.ws.max_row)
        c1.add_data(data, titles_from_data=True)

        self.ws.add_chart(c1, "F13")

    # 设置cell值的函数
    def set_value(self, row, col, value):
        print("正在设置 %s" % value)
        self.ws.cell(row=row, column=col).value = value

    def process(self):

        ws = ExcelProcess(self.source_path, self.test_device, self.source_file_name,
                          self.target_file_name)
        current = ws.getAverageValue()

        i_light = current[0]
        i_dark = current[1]
        i_photo_current = i_light - i_dark

        ws.re_draw()
        # 表头
        title_list = ["file_name", "I_light", "I_dark", "I_photo"]
        for i in range(1, len(title_list) + 1):
            title = title_list[i - 1]
            ws.set_value(row=1, col=i + 10, value=title)
        # 值
        value_list = [self.sheet, i_light, i_dark, i_photo_current]
        for i in range(1, len(value_list) + 1):
            value = value_list[i - 1]
            ws.set_value(row=2, col=i + 10, value=value)
        ws.wb.save(self.target_file_name)
        ws.wb.close()


# 主程序
class MainProcess(ExcelDir):

    def __init__(self, SourcePath, TestDevice):
        super().__init__(SourcePath, TestDevice)
        self.source_dir = ExcelDir(self.source_path, self.test_device).SourceFileList()
        self.target_dir = ExcelDir(self.source_path, self.test_device).TargetFileList()

    # 添加新参数时需要修改表头
    def merge_xlsx_file(self):
        wb = Workbook()  # 打开第一张电子表格
        ws = wb.active  # 激活 worksheet
        ws.title = 'merged result'  # 合并结果

        # 添加新参数时需要修改
        ws.cell(1, 1, "file_name")
        ws.cell(1, 2, "I_light")
        ws.cell(1, 3, "I_dark")
        ws.cell(1, 4, "I_photo")

        for filename in self.target_dir:
            work_book = load_workbook(filename)
            sheet = work_book.active  # 激活 worksheet
            for row in sheet.iter_rows(min_row=2, max_row=2, min_col=11, max_col=14):  # 从第二行开启迭代
                values = [cell.value for cell in row]  # 返回一列的值，以列表类型
                ws.append(values)  # 把列表增加到新的表格里面
        return wb

    def main(self):
        for i in range(0, len(self.source_dir)):
            ExcelProcess(self.source_path, self.test_device,
                         self.source_dir[i], self.target_dir[i]).process()
            # print("循环一次了")
        # 汇总目标文件夹的数据到 当前目录下的 merge_data.xlsx
        wb = self.merge_xlsx_file()
        name = ExcelDir(self.source_path, self.test_device).target_path + os.path.basename(
            os.path.dirname(ExcelDir(self.source_path, self.test_device).target_path)) + "_汇总.xlsx"
        wb.save(name)  # 保存数据到硬盘
        wb.close()


# 转换csv_to_Excel
class DirCsv(object):
    def __init__(self, SourcePath):
        self.path = SourcePath

    @staticmethod
    def IsSubString(SubStrList, Str):
        """
        #判断字符串Str是否包含序列SubStrList中的每一个子字符串
        #>>>SubStrList=['F','EMS','txt']
        #>>>Str='F06925EMS91.txt'
        #>>>IsSubString(SubStrList,Str)#return True (or False)
        """
        flag = True
        for substr in SubStrList:
            if not (substr in Str):
                flag = False
        return flag

    def SourceFileList(self):
        """
        #获取目录中指定的文件名
        #>>>FlagStr=['F','EMS','txt'] #要求文件名称中包含这些字符
        #>>>source_file_list=GetFileList(FindPath,FlagStr) #
        """

        FlagStr = ".csv"

        from os import listdir, path
        source_file_list = []
        FileNames = listdir(self.path)
        if len(FileNames) > 0:
            for fn in FileNames:
                if len(FlagStr) > 0:
                    # 返回指定类型的文件名
                    if self.IsSubString(FlagStr, fn):
                        full_file_name = path.join(self.path, fn)
                        source_file_list.append(full_file_name)
                else:
                    # 默认直接返回所有文件名
                    full_file_name = path.join(self.path, fn)
                    source_file_list.append(full_file_name)

        # 对文件名排序
        if len(source_file_list) > 0:
            source_file_list.sort()

        return source_file_list

    def TargetFileList(self):

        source_file_list = self.SourceFileList()
        target_file_list = []
        for dirs in source_file_list:
            directory = dirs.replace(".csv", ".xlsx", 1)
            target_file_list.append(directory)
        return target_file_list

    def csv_to_excel(self, file_name, target_name):

        data = read_csv(file_name, header=8)
        sheet_name = target_name.replace(self.path, "", 1)
        sheet_name2 = sheet_name.replace(".xlsx", "", 1)
        data.to_excel(target_name, sheet_name=sheet_name2)

    def csv_to_xlsx(self):
        SourcePath = self.path
        # SourcePath = input("请输入待转换csv路径(eg. -> './test/') : ")
        source_file_list = DirCsv(SourcePath).SourceFileList()
        target_file_name = DirCsv(SourcePath).TargetFileList()
        for i in range(0, len(source_file_list)):
            DirCsv(SourcePath).csv_to_excel(source_file_list[i], target_file_name[i])
        print("已在 %s 下将.csv 转换为 .xlsx" % SourcePath)


# 小程序文字提示
def print_test():
    print("-" * ASTERISK)
    print("I-t测试数据处理小程序 4.1")
    print("-" * ASTERISK)
    print("1. 测试文件，例如输入以下参数，表示:")
    print("原始数据在当前文件夹的test文件下")
    print("处理过的数据生成在当前文件夹的test_precessed目录下")
    print("原始数据是由 PDA 设备采集的，可选另一参数为2636B")
    print("2. 输入原文件路径eg.(./test/)，和'测试机器 'PDA' or '2636B'")
    print("3. 目录不要输入错误了，记得输入的目录以'/'结尾！")
    print("3. 目录不要输入错误了，记得输入的目录以'/'结尾！")
    print("3. 目录不要输入错误了，记得输入的目录以'/'结尾！")
    print("eg. 程序放在在'/Users/yl/Desktop/data_process/'目录下")
    print("那么可以输入绝对路径 '/Users/yl/Desktop/data_process/10-1/405/'")
    print("或者输入 './10-1/405/' (Windows可能不行，Linux系统可以输入 '.' 表示当前目录)")
    print("-" * ASTERISK)


# 套娃了
def process(source_path, test_device):
    MainProcess(source_path, test_device).main()
    target_path = ExcelDir(source_path, test_device).target_path
    print("-" * ASTERISK)
    print("处理完成！")
    print("在 %s 目录下生成了处理过的文件" % target_path)
    name = ExcelDir(source_path, test_device).target_path + os.path.basename(
        os.path.dirname(ExcelDir(source_path, test_device).target_path)) + "_汇总.xlsx"

    print("在 %s 目录下生成了 %s" % (target_path, name.rstrip("/")))


# 忍不住套娃。。
def main_main():
    """
    测试文件，例如输入以下参数，表示：
    原始数据在当前文件夹的test文件下
    处理过的数据生成在当前文件夹的test_processed目录下
    原始数据是由 PDA 设备采集的，可选另一参数为2636B
    输入原文件路径，和测试机器 PDA or 2636B
    """

    # source_path = "./10-1/405/"
    # test_device = "2636B"
    print_test()
    while True:
        source_path = input("退出请输入 'exit'\n请输入源文件夹路径：")
        if source_path != "exit":
            test_device = input("请输入测试设备（PDA or 2636B）：")
            isExists = os.path.exists(source_path)
            if isExists:
                print(source_path + "目录存在")
                target_path = os.path.dirname(source_path) + "_processed"
                mkdir(target_path)
                if test_device == "PDA":
                    process(source_path, test_device)
                    break
                elif test_device == "2636B":
                    DirCsv(source_path).csv_to_xlsx()
                    process(source_path, test_device)
                    break
                else:
                    print("参数输入错误，请重新输入...")
            else:
                # 如果目录存在则不创建，并提示目录已存在
                print('目录不存在，请重新输入')
        else:
            exit()


# 程序小循环
while True:
    main_main()
    print("-" * ASTERISK)
    order = input("退出请输入 'exit' 来退出小程序~\n继续处理数据就随便输入")
    if order == "exit":
        exit()
