# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from pandas import read_csv
from sys import exit
import os.path
from os import listdir
from excel_process import ExcelProcess
from settings import Settings


def mkdir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")

    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + " 目录已存在")
        return False


def IsSubString(SubStrList, Str):
    """
    #判断字符串Str是否包含序列SubStrList中的每一个子字符串
    #>>>SubStrList=['F','EMS','txt']
    #>>>Str='F06925EMS91.txt'
    #>>>IsSubString(SubStrList,Str)#return True (or False)
    :param SubStrList:
    :param Str:
    :return:
    """
    flag = True
    for substr in SubStrList:
        if not (substr in Str):
            flag = False
    return flag


def SourceFileList(source_path: str, FlagStr: str):
    """
    #获取目录中指定的文件名
    #>>>FlagStr=['F','EMS','txt'] #要求文件名称中包含这些字符
    #>>>source_file_list=GetFileList(FindPath,FlagStr) #
    :param source_path: path
    :param FlagStr: 'xlsx' or 'csv'
    :return:

    """
    source_file_list = []
    FileNames = listdir(source_path)
    if len(FileNames) > 0:
        for fn in FileNames:
            if len(FlagStr) > 0:
                # 返回指定类型的文件名
                if IsSubString(FlagStr, fn):
                    full_file_name = os.path.join(source_path, fn)
                    source_file_list.append(full_file_name)
            else:
                # 默认直接返回所有文件名
                full_file_name = os.path.join(source_path, fn)
                source_file_list.append(full_file_name)

    # 对文件名排序
    if len(source_file_list) > 0:
        source_file_list.sort()

    return source_file_list


def TargetFileList(source_path: str, FlagStr: str) -> list:
    """

    :param source_path: path
    :param FlagStr: 'xlsx' or 'csv'
    :return: list of target file
    """
    source_file_list = SourceFileList(source_path, FlagStr)
    target_file_list = []
    if FlagStr == 'xlsx':
        for dirs in source_file_list:
            dir_name = os.path.dirname(dirs) + '_processed'
            base_name = os.path.basename(dirs)
            target_name = os.path.join(dir_name, base_name)
            target_file_list.append(target_name)
        return target_file_list
    elif FlagStr == "csv":
        for dirs in source_file_list:
            directory = dirs.replace(".csv", ".xlsx", 1)
            target_file_list.append(directory)
        return target_file_list


def csv_to_xlsx(source_path: str):
    # 转换csv_to_Excel
    source_file_list = SourceFileList(source_path, FlagStr='csv')
    target_file_list = TargetFileList(source_path, FlagStr='csv')
    for i in range(0, len(source_file_list)):
        data = read_csv(source_file_list[i], header=8)
        sheet_name = os.path.basename(target_file_list[i])
        sheet_name2 = sheet_name.replace(".xlsx", "", 1)
        data.to_excel(target_file_list[i], sheet_name=sheet_name2)
    print("已在 %s 下将.csv 转换为 .xlsx" % source_path)


# 小程序文字提示
def print_test(asterisk):
    print("-" * asterisk)
    print("I-t测试数据处理小程序 4.2")
    print("-" * asterisk)
    print("1. 测试文件，例如输入以下参数，表示:")
    print("原始数据在当前文件夹的test文件下")
    print("处理过的数据生成在当前文件夹的test_precessed目录下")
    print("原始数据是由 PDA 设备采集的，可选另一参数为2636B")
    print("2. 输入原文件路径eg.(./test/)，和'测试机器 'PDA' or '2636B'")
    print("eg. 程序放在在'/Users/yl/Desktop/data_process/'目录下")
    print("那么可以输入绝对路径 '/Users/yl/Desktop/data_process/10-1/405/'")
    print("或者输入 './10-1/405/' (Windows可能不行，Linux系统可以输入 '.' 表示当前目录)")
    print("-" * asterisk)


# 添加新参数时需要修改表头
def merge_excel_file(target_file_list):
    wb = Workbook()  # 打开第一张电子表格
    ws = wb.active  # 激活 worksheet
    ws.title = 'merged result'  # 合并结果

    # 添加新参数时需要修改
    ws.cell(1, 1, "file_name")
    ws.cell(1, 2, "I_light")
    ws.cell(1, 3, "I_dark")
    ws.cell(1, 4, "I_photo")

    for filename in target_file_list:
        work_book = load_workbook(filename)
        sheet = work_book.active  # 激活 worksheet
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=11, max_col=14):  # 从第二行开启迭代
            values = [cell.value for cell in row]  # 返回一列的值，以列表类型
            ws.append(values)  # 把列表增加到新的表格里面
    return wb


def process_excel_files(test_device, source_file_list, target_file_list, ai_settings, name):
    for i in range(0, len(source_file_list)):
        ExcelProcess(test_device, source_file_list[i], target_file_list[i], ai_settings).process()
    # 汇总目标文件夹的数据到 当前目录下的 merge_data.xlsx
    wb = merge_excel_file(target_file_list)
    wb.save(name)  # 保存数据到硬盘
    wb.close()


def process(source_path: str, test_device):
    ai_settings = Settings()
    target_path = os.path.dirname(source_path.rstrip('/') + '/') + "_processed/"
    source_file_list = SourceFileList(source_path, FlagStr='xlsx')
    target_file_list = TargetFileList(source_path, FlagStr='xlsx')
    name = target_path + os.path.basename(os.path.dirname(target_path)) + "_汇总.xlsx"

    process_excel_files(test_device, source_file_list, target_file_list, ai_settings, name)

    print("-" * ai_settings.asterisk)
    print("处理完成！")
    print("在 %s 目录下生成了处理过的文件" % target_path)
    print("在 %s 目录下生成了 %s" % (target_path, os.path.basename(name.rstrip("/"))))


def decide_device(source_path, test_device):
    if test_device == "PDA":
        process(source_path, test_device)
    elif test_device == "2636B":
        csv_to_xlsx(source_path)
        process(source_path, test_device)


def program():
    while True:
        source_path = input("退出请输入 'q'\n请输入源文件夹路径: ").rstrip("/")
        if source_path != "q":
            isExists = os.path.exists(source_path)
            if isExists:
                test_device = input("请输入测试设备（PDA or 2636B）：")
                if test_device == "PDA" or test_device == "2636B":
                    print(source_path + " 目录存在")
                    target_path = os.path.dirname(source_path + '/') + "_processed"
                    if mkdir(target_path):
                        decide_device(source_path, test_device)
                        break
                    elif input("是否覆盖目录 %s (y/n)" % os.path.basename(target_path)) == 'y':
                        decide_device(source_path, test_device)
                        break
                    else:
                        break
                else:
                    print("测试设备输入错误，请重新输入...")
            else:
                # 如果目录存在则不创建，并提示目录已存在
                print('目录不存在，请重新输入')

        else:
            exit()
