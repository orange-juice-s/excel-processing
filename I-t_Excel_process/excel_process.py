# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from heapq import nlargest, nsmallest
from numpy import mean
from openpyxl.chart import Reference, LineChart
from pandas import read_excel


class ExcelProcess(object):

    def __init__(self, test_device, source_file_name, target_file_name, ai_settings):
        # 获取源文件目录的列表 不能在这里调用 for 否则只能读出列表最后一项的函数值
        self.test_device = test_device
        self.target_file_name = target_file_name
        self.source_file_name = source_file_name
        self.ai_settings = ai_settings
        self.wb = load_workbook(source_file_name)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        # 判断测试设备
        if self.test_device == '2636B':
            self.start_row = int(self.ws.max_row * self.ai_settings.B2636_percentage_ignored)
            self.start_col = self.ai_settings.B2636_start_col
        elif self.test_device == 'PDA':
            self.start_row = int(self.ws.max_row * self.ai_settings.PDA_percentage_ignored)
            self.start_col = self.ai_settings.PDA_start_col

    # 定义一个获取前20个最大最小值的平均数（去除了表格前10-15%的抖动区域），即获取光电流和暗电流
    def getAverageValue(self):
        print("-" * self.ai_settings.asterisk)
        print("正在获取 %s 的光暗电流值" % self.source_file_name)

        if self.test_device == "PDA":
            data = read_excel(self.source_file_name, header=None, usecols=[1], skiprows=self.start_row)
            data_abs = data.abs()
            column_data = data_abs[1].to_list()
            column_max = nlargest(self.ai_settings.photo_current_count_numbers, column_data)
            column_min = nsmallest(self.ai_settings.photo_current_count_numbers, column_data)
            # 取其平均值
            averageMaxValue = mean(column_max)
            averageMinValue = mean(column_min)
            # 返回这两个平均值（光电流和暗电流）
            return averageMaxValue, averageMinValue
        elif self.test_device == "2636B":
            data = read_excel(self.source_file_name, header=None, usecols=[3], skiprows=self.start_row)
            data_abs = data.abs()
            column_data = data_abs[3].to_list()
            column_max = nlargest(self.ai_settings.photo_current_count_numbers, column_data)
            column_min = nsmallest(self.ai_settings.photo_current_count_numbers, column_data)
            # 取其平均值
            averageMaxValue = mean(column_max)
            averageMinValue = mean(column_min)
            # 返回这两个平均值（光电流和暗电流）
            return averageMaxValue, averageMinValue

    # 在新的path中画一个简要的I-t曲线。x轴是1， 2， 3， ... 未用到真实的t时间。
    def re_draw(self):

        c1 = LineChart()
        c1.title = "2D Line Chart"
        c1.legend = None
        c1.style = 15
        c1.y_axis.title = 'Size'
        c1.x_axis.title = 'Test Number'

        data = Reference(self.ws, min_col=self.start_col, min_row=self.start_row, max_col=self.start_col,
                         max_row=self.ws.max_row)
        c1.add_data(data, titles_from_data=True)

        self.ws.add_chart(c1, "F13")

    # 设置cell值的函数
    def set_value(self, row, col, value):
        print("正在设置 %s" % value)
        self.ws.cell(row=row, column=col).value = value

    def process(self):

        ws = ExcelProcess(self.test_device, self.source_file_name,
                          self.target_file_name, self.ai_settings)
        current = ws.getAverageValue()

        i_light = current[0]
        i_dark = current[1]
        i_photo_current = i_light - i_dark

        ws.re_draw()
        # 表头
        title_list = ["file_name", "I_light", "I_dark", "I_photo"]
        for i in range(1, len(title_list) + 1):
            title = title_list[i - 1]
            ws.set_value(row=1, col=i + 10, value=title)
        # 值
        value_list = [self.sheet, i_light, i_dark, i_photo_current]
        for i in range(1, len(value_list) + 1):
            value = value_list[i - 1]
            ws.set_value(row=2, col=i + 10, value=value)
        ws.wb.save(self.target_file_name)
        ws.wb.close()
